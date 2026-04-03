from io import BytesIO
from datetime import date
from django.contrib.auth import get_user_model

User = get_user_model()


def generate_monthly_excel(user, year: int, month: int) -> bytes:
    """Erstellt einen Excel-Monatsbericht mit 4 Sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from apps.timesessions.models import TimeEntry
    from apps.absences.models import AbsenceRequest
    from apps.overtime.models import OvertimeTransaction
    from apps.overtime.services import OvertimeCalculator

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Zeiteinträge
    ws1 = wb.active
    ws1.title = "Zeiteinträge"
    style_header(ws1, ["Datum", "Wochentag", "Beginn", "Ende", "Pause (min)", "Netto (h)", "Status", "Anmerkung"])
    entries = TimeEntry.objects.filter(
        user=user, date__year=year, date__month=month,
        status__in=["COMPLETED", "AUTO_CLOSED", "MANUAL"],
    ).order_by("date")
    weekdays = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    for row, e in enumerate(entries, 2):
        ws1.append([
            e.date.strftime("%d.%m.%Y"),
            weekdays[e.date.weekday()],
            e.start_time.strftime("%H:%M") if e.start_time else "",
            e.end_time.strftime("%H:%M") if e.end_time else "",
            e.break_minutes,
            round(e.net_minutes / 60, 2),
            e.get_status_display(),
            e.notes,
        ])

    # Sheet 2: Abwesenheiten
    ws2 = wb.create_sheet("Abwesenheiten")
    style_header(ws2, ["Von", "Bis", "Typ", "Tage", "Status", "AU liegt vor", "AU eingereicht am"])
    absences = AbsenceRequest.objects.filter(
        user=user, start_date__year=year, status="APPROVED"
    ).select_related("leave_type")
    for a in absences:
        is_sick = a.leave_type.code == "SICK"
        ws2.append([
            a.start_date.strftime("%d.%m.%Y"),
            a.end_date.strftime("%d.%m.%Y"),
            a.leave_type.name,
            float(a.duration_days or 0),
            a.get_status_display(),
            ("Ja" if a.au_vorhanden else "Nein") if is_sick else "–",
            a.au_eingereicht_am.strftime("%d.%m.%Y") if is_sick and a.au_eingereicht_am else "–",
        ])

    # Sheet 3: Urlaubskonto
    ws3 = wb.create_sheet("Urlaubskonto")
    style_header(ws3, ["Anspruch", "Verbraucht", "Rest"])
    profile = getattr(user, "userprofile", None)
    annual = float(profile.annual_leave_days) if profile else 30
    used = float(AbsenceRequest.objects.filter(
        user=user, leave_type__code="VACATION", status="APPROVED", start_date__year=year
    ).aggregate(t=__import__("django.db.models", fromlist=["Sum"]).Sum("duration_days"))["t"] or 0)
    ws3.append([annual, used, annual - used])

    # Sheet 4: Überstundenkonto
    ws4 = wb.create_sheet("Überstundenkonto")
    style_header(ws4, ["Datum", "Typ", "Minuten", "Stunden", "Monat", "Grund"])
    ot_txs = OvertimeTransaction.objects.filter(
        account__user=user, transaction_date__year=year
    ).order_by("transaction_date")
    for tx in ot_txs:
        ws4.append([
            tx.transaction_date.strftime("%d.%m.%Y"),
            tx.get_transaction_type_display(),
            tx.amount_minutes,
            round(tx.amount_minutes / 60, 2),
            tx.reference_month,
            tx.reason,
        ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
